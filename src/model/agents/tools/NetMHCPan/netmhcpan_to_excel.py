import re
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

def save_excel(output:str,output_dir:str,output_filename:str):
    table_pattern = re.compile(r"(\d+)\s+([^\s]+)\s+([A-Z]+)\s+([A-Z]+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+([A-Z]+)\s+([^\s]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([<= WS B]*)")
    matches = table_pattern.findall(output)

    # 将匹配的数据转换为 DataFrame
    columns = ["Pos", "MHC", "Peptide", "Core", "Of", "Gp", "Gl", "Ip", "Il", "Icore", "Identity", "Score_EL", "%Rank_EL", "Score_BA", "%Rank_BA", "Aff(nM)", "BindLevel"]
    df = pd.DataFrame(matches, columns=columns)

    # 提取包含 Allele 的整行数据
    summary_pattern = re.compile(r".*Allele\s+[^\s]+.*")
    summary_match = summary_pattern.findall(output)

    # 如果找到匹配的统计信息，将其添加到 Results 表的最后一行
    if summary_match:
        # 创建一个新行，前 17 列合并为一个单元格
        summary_row = [summary_match[0]] + [""] * (len(columns) - 1)  # 第一列是 summary_match[0]，其余列为空

        # 将 summary_row 添加到 DataFrame 的最后一行
        df.loc[len(df)] = summary_row
        
    output_path= Path(output_dir) / output_filename

    # 写入 Excel 文件
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Results", index=False)

        # 获取 workbook 和 worksheet 对象
        workbook = writer.book
        worksheet = writer.sheets["Results"]

        # 如果存在 summary_match，合并最后一行的前 17 列并居中
        if summary_match:
            # 合并最后一行的前 17 列
            worksheet.merge_cells(start_row=len(df)+1, start_column=1, end_row=len(df)+1, end_column=17)

            # 设置合并后的单元格内容居中
            cell = worksheet.cell(row=len(df)+1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存文件
    workbook.save(output_path)